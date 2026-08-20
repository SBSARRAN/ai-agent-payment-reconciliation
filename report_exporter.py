
# IMPORTS


# Path is used to handle the report file path.
from pathlib import Path

# openpyxl is used to create and update Excel files.
from openpyxl import Workbook, load_workbook



# REPORT FILE


REPORT_PATH = Path(
    "data/reports/reconciliation_report.xlsx"
)



# EXPORT MATCHED PAYMENT


def export_matched_payment(receipt, matched_transaction):

    # Create reports folder if it does not exist.
    REPORT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True
    )


    
    # IF REPORT ALREADY EXISTS
    

    if REPORT_PATH.exists():

        # Open existing Excel file.
        workbook = load_workbook(
            REPORT_PATH
        )

        worksheet = workbook.active


    
    # IF REPORT DOES NOT EXIST
    

    else:

        # Create a new Excel workbook.
        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Matched Payments"


        # Create Excel column headers.
        worksheet.append([
            "Customer Name",
            "Phone Number",
            "Amount",
            "UTR",
            "Transaction ID",
            "Payment Date",
            "Payment Time",
            "Payment App",
            "Receipt Payer",
            "Receipt Payee",
            "Statement Payer",
            "Match Status"
        ])


    
    # ADD MATCHED PAYMENT ROW
    

    worksheet.append([
        receipt.customer_name,
        receipt.phone_number,
        receipt.amount,
        receipt.utr,
        receipt.transaction_id,
        receipt.payment_date,
        receipt.payment_time,
        receipt.payment_app,
        receipt.payer_name,
        receipt.payee_name,
        matched_transaction.payer_name,
        "MATCHED"
    ])


    
    # SAVE EXCEL FILE
    

    workbook.save(
        REPORT_PATH
    )


    print(
        f"REPORT UPDATED: {REPORT_PATH}"
    )